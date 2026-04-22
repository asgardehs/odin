"""Parse an XLSX workbook into headers + rows, emit JSON on stdout.

Usage:
    python parse_xlsx.py <path/to/file.xlsx>

Output shape (stdout):
    {
        "headers": ["First Name", "Last Name", ...],
        "rows":    [{"First Name": "Alice", ...}, ...],
        "sheet":   "Sheet1",
        "sheet_count": 1
    }

The first sheet is always used. If the workbook has more than one sheet a
warning is written to stderr (but parsing still succeeds).

openpyxl is required at runtime; it is installed once into the odin cache
directory by the Go wrapper (internal/ratatoskr/ratatoskr.go) and the path
is handed to this script via PYTHONPATH.
"""

import json
import sys

try:
    from openpyxl import load_workbook
except ImportError as exc:
    print(json.dumps({"error": f"openpyxl import failed: {exc}"}), file=sys.stderr)
    sys.exit(2)


def cell_value(v):
    """Coerce an openpyxl cell value to a JSON-safe string.

    openpyxl returns Python primitives: str, int, float, bool, datetime,
    and None. Dates/datetimes must round-trip through ISO-8601 so downstream
    mappers can reuse their CSV date parsers.
    """
    if v is None:
        return ""
    # datetime.datetime / datetime.date — isoformat() handles both.
    if hasattr(v, "isoformat"):
        return v.isoformat()
    if isinstance(v, bool):
        return "true" if v else "false"
    if isinstance(v, float):
        # Avoid noisy 3.0 → "3.0" when the value is really an integer.
        if v.is_integer():
            return str(int(v))
        return repr(v)
    return str(v)


def main():
    if len(sys.argv) != 2:
        print(json.dumps({"error": "usage: parse_xlsx.py <path>"}), file=sys.stderr)
        sys.exit(2)
    path = sys.argv[1]

    try:
        # data_only=True resolves formula cells to their cached values.
        # read_only=True streams sheets without holding the whole workbook in
        # memory (important for 10k-row files).
        wb = load_workbook(filename=path, data_only=True, read_only=True)
    except Exception as exc:  # noqa: BLE001 — stringify any openpyxl failure
        print(json.dumps({"error": f"load_workbook: {exc}"}), file=sys.stderr)
        sys.exit(1)

    sheet_names = wb.sheetnames
    if not sheet_names:
        print(json.dumps({"error": "workbook has no sheets"}), file=sys.stderr)
        sys.exit(1)

    if len(sheet_names) > 1:
        print(
            f"warning: workbook has {len(sheet_names)} sheets; using first ({sheet_names[0]!r})",
            file=sys.stderr,
        )

    ws = wb[sheet_names[0]]

    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        print(json.dumps({"error": "sheet has no header row"}), file=sys.stderr)
        sys.exit(1)

    headers = []
    for cell in header_row:
        headers.append(cell_value(cell).strip())

    # Strip trailing blank header columns (common when openpyxl reports the
    # sheet's max column as one past the actual data).
    while headers and headers[-1] == "":
        headers.pop()

    rows = []
    for rec in rows_iter:
        if rec is None:
            continue
        # Skip rows where every cell is empty — openpyxl can emit padding
        # rows after the real data in streaming mode.
        cells = [cell_value(c) for c in rec[: len(headers)]]
        if not any(cells):
            continue
        row = {}
        for i, h in enumerate(headers):
            if h == "":
                continue
            if i < len(cells):
                row[h] = cells[i].strip()
            else:
                row[h] = ""
        rows.append(row)

    out = {
        "headers": headers,
        "rows": rows,
        "sheet": sheet_names[0],
        "sheet_count": len(sheet_names),
    }
    json.dump(out, sys.stdout)


if __name__ == "__main__":
    main()
